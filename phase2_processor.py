"""
PDF Invoice Processor - Phase 2: Classification, Grouping & Extraction

Reads the manifest.json from Phase 1, uses Ollama with a VISION model
(qwen3-vl:8b) to analyze page images directly:
  1. Classify each page by looking at its image
  2. Group pages into invoices (text-based, using classification results)
  3. Extract structured fields by looking at invoice page images
  4. Output results as a CSV/XLSX spreadsheet

The vision model sees the actual document layout — tables, headers, logos,
spatial positioning — rather than relying on OCR text alone. OCR text from
Phase 1 is optionally included as supplementary context.

All processing is local via Ollama — no data leaves the machine.

Requirements:
    - Phase 1 output (manifest.json + processed PDFs + images/)
    - Ollama running locally with qwen3-vl:8b model
    - Python: openpyxl, pdfplumber, Pillow
"""

import os
import sys
import csv
import json
import logging
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing 'openpyxl'. Run: pip install openpyxl")

try:
    import pdfplumber
except ImportError:
    sys.exit("Missing 'pdfplumber'. Run: pip install pdfplumber")

from ollama_client import OllamaClient, OllamaError
from prompts import (
    SYSTEM_PROMPT,
    CLASSIFY_PAGE_PROMPT,
    CLASSIFY_PAGE_WITH_TEXT_PROMPT,
    GROUP_PAGES_PROMPT,
    EXTRACT_FIELDS_PROMPT,
    EXTRACT_FIELDS_WITH_TEXT_PROMPT,
    EXTRACT_KEY_FIELDS_PROMPT,
)

LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
MAX_TEXT_LENGTH = 4000  # truncate supplementary OCR text
MAX_IMAGES_PER_EXTRACTION = 4  # max pages to send as images in one call
LOW_CONFIDENCE_RETRY_THRESHOLD = 0.6


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class PageClassification:
    page_id: str
    source_pdf: str
    original_page: int
    page_type: str
    confidence: float
    invoice_number: Optional[str]
    page_indicator: Optional[str]
    reasoning: str
    image_path: str


@dataclass
class InvoiceGroup:
    group_id: int
    group_type: str
    invoice_number: Optional[str]
    page_ids: list
    pages: list


@dataclass
class InvoiceRecord:
    group_id: int
    page_ids: list
    source_files: list
    invoice_number: Optional[str]
    invoice_date: Optional[str]
    due_date: Optional[str]
    supplier_name: Optional[str]
    supplier_address: Optional[str]
    customer_name: Optional[str]
    customer_address: Optional[str]
    purchase_order_number: Optional[str]
    currency: Optional[str]
    line_items_summary: str
    subtotal: Optional[float]
    shipping: Optional[float]
    gst: Optional[float]
    pst: Optional[float]
    hst: Optional[float]
    other_taxes: Optional[float]
    total: Optional[float]
    notes: Optional[str]
    extraction_confidence: float
    fields_uncertain: list
    num_pages: int
    group_type: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_page_text(pdf_path: str) -> str:
    """Extract text from a single-page PDF (supplementary to vision)."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if pdf.pages:
                return (pdf.pages[0].extract_text() or "").strip()
    except Exception:
        pass
    return ""


def truncate_text(text: str, max_len: int = MAX_TEXT_LENGTH) -> str:
    if len(text) <= max_len:
        return text
    half = max_len // 2
    return text[:half] + "\n\n[... truncated ...]\n\n" + text[-half:]


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _first_non_empty(*vals):
    for val in vals:
        if val is None:
            continue
        if isinstance(val, str):
            if val.strip():
                return val.strip()
        else:
            return val
    return None


def _build_extraction_prompt(base_prompt: str, ordered_page_ids: list[str]) -> str:
    if len(ordered_page_ids) <= 1:
        return base_prompt
    page_order_lines = "\n".join(
        f"- Image {idx + 1}: {page_id}" for idx, page_id in enumerate(ordered_page_ids)
    )
    return (
        f"{base_prompt}\n\n"
        "Important multi-page guidance:\n"
        "The images are consecutive invoice pages in this exact order:\n"
        f"{page_order_lines}\n"
        "- Use page 1 for header fields (supplier/customer, invoice number/date).\n"
        "- Use later pages for continued line items, taxes, and totals.\n"
        "- Return a single consolidated JSON for the whole invoice."
    )


def _merge_two_page_rescue(primary: dict, page1: dict, page2: dict) -> dict:
    merged = dict(primary)

    text_fields = [
        "invoice_number", "invoice_date", "due_date",
        "supplier_name", "supplier_address",
        "customer_name", "customer_address",
        "purchase_order_number", "currency", "notes",
    ]
    for field in text_fields:
        merged[field] = _first_non_empty(
            primary.get(field), page1.get(field), page2.get(field)
        )

    amount_fields = ["subtotal", "shipping", "gst", "pst", "hst", "other_taxes", "total"]
    for field in amount_fields:
        merged[field] = _first_non_empty(
            _to_float(primary.get(field)),
            _to_float(page2.get(field)),
            _to_float(page1.get(field)),
        )

    if not isinstance(merged.get("line_items"), list):
        merged["line_items"] = []

    uncertain = merged.get("fields_uncertain", [])
    if not isinstance(uncertain, list):
        uncertain = []
    resolved = set()
    for field in text_fields + amount_fields:
        val = merged.get(field)
        if val is not None and (not isinstance(val, str) or val.strip()):
            resolved.add(field)
    merged["fields_uncertain"] = [f for f in uncertain if f not in resolved]

    primary_conf = float(primary.get("extraction_confidence", 0.0) or 0.0)
    page1_conf = float(page1.get("extraction_confidence", 0.0) or 0.0)
    page2_conf = float(page2.get("extraction_confidence", 0.0) or 0.0)
    merged["extraction_confidence"] = max(
        primary_conf,
        min(0.85, (primary_conf * 0.5) + (page1_conf * 0.25) + (page2_conf * 0.25) + 0.1),
    )
    return merged


# ---------------------------------------------------------------------------
# Step 1: Classify pages using VISION
# ---------------------------------------------------------------------------
def classify_pages(
    manifest: dict,
    client: OllamaClient,
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> list[PageClassification]:
    """Classify each page by sending its image to the vision model."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    pages = manifest.get("pages", [])
    total = len(pages)
    classifications = []

    log(f"Classifying {total} pages using vision model...")

    for i, page_data in enumerate(pages):
        if cancel_event and cancel_event.is_set():
            break

        page_id = page_data["page_id"]
        image_path = page_data.get("output_image", "")
        pdf_path = page_data.get("output_pdf", "")

        if progress_callback:
            progress_callback(i + 1, total, f"Classifying {page_id}")

        # Check image exists
        if not image_path or not os.path.exists(image_path):
            log(f"  {page_id}: No image found — classifying as unknown")
            classifications.append(PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type="unknown", confidence=0.0,
                invoice_number=None, page_indicator=None,
                reasoning="No image available", image_path="",
            ))
            continue

        # Build prompt — optionally include OCR text as supplementary context
        if include_ocr_text and pdf_path and os.path.exists(pdf_path):
            ocr_text = get_page_text(pdf_path)
            if len(ocr_text) > 20:
                prompt = CLASSIFY_PAGE_WITH_TEXT_PROMPT.format(
                    page_text=truncate_text(ocr_text)
                )
            else:
                prompt = CLASSIFY_PAGE_PROMPT
        else:
            prompt = CLASSIFY_PAGE_PROMPT

        try:
            result = client.generate_json_with_image(
                prompt=prompt,
                image_paths=[image_path],
                system=SYSTEM_PROMPT,
                timeout=120,
            )

            cls = PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type=result.get("page_type", "unknown"),
                confidence=float(result.get("confidence", 0.0)),
                invoice_number=result.get("invoice_number"),
                page_indicator=result.get("page_indicator"),
                reasoning=result.get("reasoning", ""),
                image_path=image_path,
            )
            classifications.append(cls)

            inv = f" INV#{cls.invoice_number}" if cls.invoice_number else ""
            log(f"  {page_id}: {cls.page_type} (conf:{cls.confidence:.2f}){inv}")

        except OllamaError as e:
            log(f"  {page_id}: ERROR — {e}")
            classifications.append(PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type="unknown", confidence=0.0,
                invoice_number=None, page_indicator=None,
                reasoning=f"VLM error: {e}", image_path=image_path,
            ))

    return classifications


# ---------------------------------------------------------------------------
# Step 2: Group pages (text-based — uses classification metadata)
# ---------------------------------------------------------------------------
def group_pages(
    classifications: list[PageClassification],
    client: OllamaClient,
    log_callback=None,
) -> list[InvoiceGroup]:
    """Group classified pages into invoices. This step is text-based (no images)."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    pages_summary = []
    for cls in classifications:
        pages_summary.append({
            "page_id": cls.page_id,
            "page_type": cls.page_type,
            "invoice_number": cls.invoice_number,
            "page_indicator": cls.page_indicator,
            "source_pdf": cls.source_pdf,
            "original_page": cls.original_page,
        })

    chunk_size = 15
    all_groups = []
    group_id_counter = 1

    for start in range(0, len(pages_summary), chunk_size):
        chunk = pages_summary[start:start + chunk_size]
        log(f"Grouping pages {start + 1}-{start + len(chunk)} of {len(pages_summary)}...")

        prompt = GROUP_PAGES_PROMPT.format(pages_json=json.dumps(chunk, indent=2))

        success = False
        for attempt in range(2):  # one retry before fallback
            try:
                # Grouping is text-only — no images needed
                result = client.generate_json(prompt, system=SYSTEM_PROMPT, timeout=600)
                raw_groups = result.get("groups", [])

                for g in raw_groups:
                    page_ids = g.get("page_ids", [])
                    group_pages_list = [c for c in classifications if c.page_id in page_ids]
                    all_groups.append(InvoiceGroup(
                        group_id=group_id_counter,
                        group_type=g.get("group_type", "invoice"),
                        invoice_number=g.get("invoice_number"),
                        page_ids=page_ids,
                        pages=group_pages_list,
                    ))
                    group_id_counter += 1

                success = True
                break

            except OllamaError as e:
                if attempt == 0:
                    log(f"  Grouping attempt {attempt + 1} failed ({e}), retrying...")
                else:
                    log(f"  Grouping error: {e}")
                    log("  Falling back to sequential grouping...")

        if not success:
            all_groups.extend(
                _fallback_grouping(classifications[start:start + chunk_size], group_id_counter)
            )
            group_id_counter = max((g.group_id for g in all_groups), default=0) + 1

    # Catch ungrouped pages
    grouped_ids = set()
    for g in all_groups:
        grouped_ids.update(g.page_ids)

    ungrouped = [c for c in classifications if c.page_id not in grouped_ids]
    if ungrouped:
        log(f"  {len(ungrouped)} pages not grouped — adding individually")
        for cls in ungrouped:
            all_groups.append(InvoiceGroup(
                group_id=group_id_counter,
                group_type="invoice" if "invoice" in cls.page_type else "supporting_documents",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            ))
            group_id_counter += 1

    log(f"Created {len(all_groups)} groups")
    for g in all_groups:
        inv = f" INV#{g.invoice_number}" if g.invoice_number else ""
        log(f"  Group {g.group_id}: {g.group_type}{inv} — {len(g.page_ids)} page(s)")

    return all_groups


def _fallback_grouping(classifications, start_id):
    """Simple fallback: each first_page starts a group, continuations follow."""
    groups = []
    current = None
    gid = start_id

    for cls in classifications:
        if cls.page_type == "invoice_first_page":
            if current:
                groups.append(current)
            current = InvoiceGroup(
                group_id=gid, group_type="invoice",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            )
            gid += 1
        elif cls.page_type == "invoice_continuation" and current:
            current.page_ids.append(cls.page_id)
            current.pages.append(cls)
        else:
            if current:
                groups.append(current)
                current = None
            groups.append(InvoiceGroup(
                group_id=gid,
                group_type="supporting_documents" if cls.page_type == "supporting_document" else "unknown",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            ))
            gid += 1

    if current:
        groups.append(current)
    return groups


# ---------------------------------------------------------------------------
# Step 3: Extract fields using VISION
# ---------------------------------------------------------------------------
def extract_invoice_fields(
    groups: list[InvoiceGroup],
    manifest: dict,
    client: OllamaClient,
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> list[InvoiceRecord]:
    """Extract structured fields by sending invoice page images to the vision model."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    # Build lookups from page_id to image/pdf paths
    page_image_map = {}
    page_pdf_map = {}
    for p in manifest.get("pages", []):
        page_image_map[p["page_id"]] = p.get("output_image", "")
        page_pdf_map[p["page_id"]] = p.get("output_pdf", "")
    page_order_map = {p["page_id"]: i for i, p in enumerate(manifest.get("pages", []))}

    records = []
    invoice_groups = [g for g in groups if g.group_type == "invoice"]
    total = len(invoice_groups)

    log(f"\nExtracting fields from {total} invoice group(s) using vision model...")

    for i, group in enumerate(invoice_groups):
        if cancel_event and cancel_event.is_set():
            break

        if progress_callback:
            progress_callback(i + 1, total, f"Extracting group {group.group_id}")

        inv_label = f"Group {group.group_id}"
        if group.invoice_number:
            inv_label += f" (INV#{group.invoice_number})"

        # Collect image paths for all pages in this group
        image_paths = []
        source_files = set()
        ocr_texts = []

        ordered_page_ids = sorted(
            group.page_ids,
            key=lambda pid: page_order_map.get(pid, 10**9),
        )

        for page_id in ordered_page_ids:
            img = page_image_map.get(page_id, "")
            if img and os.path.exists(img):
                image_paths.append(img)
            pdf = page_pdf_map.get(page_id, "")
            if include_ocr_text and pdf and os.path.exists(pdf):
                text = get_page_text(pdf)
                if text:
                    ocr_texts.append(f"--- {page_id} ---\n{text}")
            for p in group.pages:
                if p.page_id == page_id:
                    source_files.add(p.source_pdf)

        if not image_paths:
            log(f"  {inv_label}: No images — skipping")
            continue

        # Limit images per call to avoid VRAM issues
        # For invoices with many pages, send first N images
        if len(image_paths) > MAX_IMAGES_PER_EXTRACTION:
            log(f"  {inv_label}: {len(image_paths)} pages, sending first {MAX_IMAGES_PER_EXTRACTION}")
            image_paths = image_paths[:MAX_IMAGES_PER_EXTRACTION]

        # Build prompt — include OCR text as supplementary context if available
        combined_ocr = "\n".join(ocr_texts)
        if combined_ocr and len(combined_ocr) > 30:
            base_prompt = EXTRACT_FIELDS_WITH_TEXT_PROMPT.format(
                invoice_text=truncate_text(combined_ocr, 6000)
            )
        else:
            base_prompt = EXTRACT_FIELDS_PROMPT
        prompt = _build_extraction_prompt(base_prompt, ordered_page_ids)

        try:
            result = client.generate_json_with_image(
                prompt=prompt,
                image_paths=image_paths,
                system=SYSTEM_PROMPT,
                timeout=240,
            )
        except OllamaError as e:
            log(f"  {inv_label}: Full extraction failed ({e}), trying simplified...")
            try:
                result = client.generate_json_with_image(
                    prompt=EXTRACT_KEY_FIELDS_PROMPT,
                    image_paths=image_paths[:1],  # just first page
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                result.setdefault("extraction_confidence", 0.3)
                result.setdefault("fields_uncertain", ["most fields"])
            except OllamaError as e2:
                log(f"  {inv_label}: All extraction failed — {e2}")
                continue

        result_conf = float(result.get("extraction_confidence", 0.0) or 0.0)
        if len(image_paths) == 2 and result_conf < LOW_CONFIDENCE_RETRY_THRESHOLD:
            log(f"  {inv_label}: Low confidence ({result_conf:.2f}) - running page-wise rescue")
            try:
                page1_result = client.generate_json_with_image(
                    prompt=EXTRACT_KEY_FIELDS_PROMPT,
                    image_paths=[image_paths[0]],
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                page2_result = client.generate_json_with_image(
                    prompt=(
                        EXTRACT_KEY_FIELDS_PROMPT
                        + "\n\nPrioritize taxes and totals if they appear on this page."
                    ),
                    image_paths=[image_paths[1]],
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                result = _merge_two_page_rescue(result, page1_result, page2_result)
            except OllamaError as rescue_error:
                log(f"  {inv_label}: Rescue pass failed - {rescue_error}")

        # Summarize line items
        line_items = result.get("line_items", [])
        if isinstance(line_items, list) and line_items:
            items_summary = "; ".join(
                item.get("description", "?") for item in line_items[:10]
            )
            if len(line_items) > 10:
                items_summary += f" (+{len(line_items) - 10} more)"
        else:
            items_summary = ""

        record = InvoiceRecord(
            group_id=group.group_id,
            page_ids=ordered_page_ids,
            source_files=sorted(source_files),
            invoice_number=result.get("invoice_number") or group.invoice_number,
            invoice_date=result.get("invoice_date"),
            due_date=result.get("due_date"),
            supplier_name=result.get("supplier_name"),
            supplier_address=result.get("supplier_address"),
            customer_name=result.get("customer_name"),
            customer_address=result.get("customer_address"),
            purchase_order_number=result.get("purchase_order_number"),
            currency=result.get("currency"),
            line_items_summary=items_summary,
            subtotal=_to_float(result.get("subtotal")),
            shipping=_to_float(result.get("shipping")),
            gst=_to_float(result.get("gst")),
            pst=_to_float(result.get("pst")),
            hst=_to_float(result.get("hst")),
            other_taxes=_to_float(result.get("other_taxes")),
            total=_to_float(result.get("total")),
            notes=result.get("notes"),
            extraction_confidence=float(result.get("extraction_confidence", 0.0)),
            fields_uncertain=result.get("fields_uncertain", []),
            num_pages=len(ordered_page_ids),
            group_type=group.group_type,
        )
        records.append(record)

        conf_pct = f"{record.extraction_confidence * 100:.0f}%"
        log(f"  {inv_label}: {record.supplier_name or '?'} → ${record.total or '?'} "
            f"(conf:{conf_pct})")

    # Also add supporting document groups
    for group in groups:
        if group.group_type != "invoice":
            records.append(InvoiceRecord(
                group_id=group.group_id,
                page_ids=group.page_ids,
                source_files=sorted(set(p.source_pdf for p in group.pages)),
                invoice_number=None,
                invoice_date=None, due_date=None,
                supplier_name=None, supplier_address=None,
                customer_name=None, customer_address=None,
                purchase_order_number=None, currency=None,
                line_items_summary="",
                subtotal=None, shipping=None,
                gst=None, pst=None, hst=None, other_taxes=None, total=None,
                notes=f"Supporting document ({len(group.page_ids)} page(s))",
                extraction_confidence=0.0, fields_uncertain=[],
                num_pages=len(group.page_ids), group_type=group.group_type,
            ))

    return records


# ---------------------------------------------------------------------------
# Output: CSV
# ---------------------------------------------------------------------------
CSV_COLUMNS = [
    "group_id", "group_type", "invoice_number", "invoice_date", "due_date",
    "supplier_name", "supplier_address", "customer_name", "customer_address",
    "purchase_order_number", "currency", "line_items_summary",
    "subtotal", "shipping", "gst", "pst", "hst", "other_taxes", "total",
    "notes", "extraction_confidence", "fields_uncertain",
    "num_pages", "page_ids", "source_files",
]


def write_csv(records: list[InvoiceRecord], output_path: str):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for rec in records:
            row = asdict(rec)
            row["page_ids"] = "; ".join(row["page_ids"])
            row["source_files"] = "; ".join(row["source_files"])
            row["fields_uncertain"] = "; ".join(row["fields_uncertain"])
            writer.writerow({k: row.get(k) for k in CSV_COLUMNS})


# ---------------------------------------------------------------------------
# Output: XLSX
# ---------------------------------------------------------------------------
def write_xlsx(records: list[InvoiceRecord], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00'
    pct_fmt = '0%'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_config = [
        ("Group", 8, None),
        ("Type", 14, None),
        ("Invoice #", 16, None),
        ("Invoice Date", 14, None),
        ("Due Date", 14, None),
        ("Supplier", 25, None),
        ("Supplier Address", 30, None),
        ("Customer", 25, None),
        ("Customer Address", 30, None),
        ("PO #", 14, None),
        ("Currency", 10, None),
        ("Line Items", 45, None),
        ("Subtotal", 14, money_fmt),
        ("Shipping", 12, money_fmt),
        ("GST", 12, money_fmt),
        ("PST", 12, money_fmt),
        ("HST", 12, money_fmt),
        ("Other Tax", 12, money_fmt),
        ("Total", 14, money_fmt),
        ("Notes", 35, None),
        ("Confidence", 12, pct_fmt),
        ("Uncertain Fields", 25, None),
        ("Pages", 8, None),
        ("Page IDs", 35, None),
        ("Source Files", 30, None),
    ]

    for col_idx, (label, width, _) in enumerate(col_config, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "A2"

    invoice_fill = PatternFill("solid", fgColor="FFFFFF")
    support_fill = PatternFill("solid", fgColor="F2F2F2")
    low_conf_fill = PatternFill("solid", fgColor="FFF2CC")

    for row_idx, rec in enumerate(records, start=2):
        row_data = [
            rec.group_id, rec.group_type, rec.invoice_number,
            rec.invoice_date, rec.due_date,
            rec.supplier_name, rec.supplier_address,
            rec.customer_name, rec.customer_address,
            rec.purchase_order_number, rec.currency,
            rec.line_items_summary,
            rec.subtotal, rec.shipping,
            rec.gst, rec.pst, rec.hst, rec.other_taxes, rec.total,
            rec.notes, rec.extraction_confidence,
            "; ".join(rec.fields_uncertain) if rec.fields_uncertain else "",
            rec.num_pages,
            "; ".join(rec.page_ids),
            "; ".join(rec.source_files),
        ]

        bg = support_fill if rec.group_type != "invoice" else invoice_fill
        if rec.group_type == "invoice" and rec.extraction_confidence < 0.5:
            bg = low_conf_fill

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.fill = bg
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            _, _, num_fmt = col_config[col_idx - 1]
            if num_fmt and value is not None:
                cell.number_format = num_fmt

    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(col_config)).column_letter}{len(records) + 1}"
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Full Phase 2 pipeline
# ---------------------------------------------------------------------------
def run_phase2(
    manifest_path: str,
    output_dir: str,
    ollama_url: str = "http://192.168.1.101:11434",
    model: str = "qwen3-vl:8b",
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> Optional[str]:
    """Run the full Phase 2 pipeline. Returns path to output XLSX or None."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    log(f"Loading manifest: {manifest_path}")
    with open(manifest_path, "r") as f:
        manifest = json.load(f)

    total_pages = len(manifest.get("pages", []))
    log(f"Manifest has {total_pages} pages from {manifest.get('total_source_files', '?')} file(s)")

    # Verify images exist
    images_found = sum(
        1 for p in manifest.get("pages", [])
        if os.path.exists(p.get("output_image", ""))
    )
    log(f"Page images found: {images_found}/{total_pages}")

    if images_found == 0:
        log("ERROR: No page images found. Did Phase 1 run successfully?")
        return None

    # Connect to Ollama
    client = OllamaClient(base_url=ollama_url, model=model)
    log(f"Connecting to Ollama at {ollama_url} with model {model}...")

    if not client.is_available():
        available = client.list_models()
        if available:
            log(f"ERROR: Model '{model}' not found. Available: {', '.join(available)}")
        else:
            log(f"ERROR: Cannot reach Ollama at {ollama_url}. Is it running?")
            log("  Start with: ollama serve")
            log(f"  Pull model: ollama pull {model}")
        return None

    log(f"Ollama connected — vision model '{model}' ready")
    ocr_mode = "with OCR text supplement" if include_ocr_text else "vision only"
    log(f"Mode: {ocr_mode}")

    os.makedirs(output_dir, exist_ok=True)

    # Step 1: Classify
    log(f"\n{'=' * 60}")
    log("STEP 1: Page Classification (vision)")
    log(f"{'=' * 60}")
    classifications = classify_pages(
        manifest, client,
        include_ocr_text=include_ocr_text,
        log_callback=log_callback,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
    )
    if cancel_event and cancel_event.is_set():
        return None

    cls_path = os.path.join(output_dir, "classifications.json")
    with open(cls_path, "w") as f:
        json.dump([asdict(c) for c in classifications], f, indent=2)
    log(f"Saved → {cls_path}")

    # Step 2: Group
    log(f"\n{'=' * 60}")
    log("STEP 2: Page Grouping")
    log(f"{'=' * 60}")
    groups = group_pages(classifications, client, log_callback=log_callback)
    if cancel_event and cancel_event.is_set():
        return None

    groups_path = os.path.join(output_dir, "groups.json")
    groups_data = []
    for g in groups:
        gd = asdict(g)
        gd.pop("pages", None)
        groups_data.append(gd)
    with open(groups_path, "w") as f:
        json.dump(groups_data, f, indent=2)
    log(f"Saved → {groups_path}")

    # Step 3: Extract
    log(f"\n{'=' * 60}")
    log("STEP 3: Field Extraction (vision)")
    log(f"{'=' * 60}")
    records = extract_invoice_fields(
        groups, manifest, client,
        include_ocr_text=include_ocr_text,
        log_callback=log_callback,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
    )
    if cancel_event and cancel_event.is_set():
        return None

    # Write outputs
    log(f"\n{'=' * 60}")
    log("OUTPUT")
    log(f"{'=' * 60}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    csv_path = os.path.join(output_dir, f"invoices_{timestamp}.csv")
    write_csv(records, csv_path)
    log(f"CSV  → {csv_path}")

    xlsx_path = os.path.join(output_dir, f"invoices_{timestamp}.xlsx")
    write_xlsx(records, xlsx_path)
    log(f"XLSX → {xlsx_path}")

    json_path = os.path.join(output_dir, f"invoices_{timestamp}.json")
    with open(json_path, "w") as f:
        json.dump([asdict(r) for r in records], f, indent=2, default=str)
    log(f"JSON → {json_path}")

    # Summary
    invoice_count = sum(1 for r in records if r.group_type == "invoice")
    support_count = sum(1 for r in records if r.group_type != "invoice")
    high_conf = sum(1 for r in records if r.group_type == "invoice" and r.extraction_confidence >= 0.7)
    low_conf = sum(1 for r in records if r.group_type == "invoice" and r.extraction_confidence < 0.5)

    log(f"\n{'=' * 60}")
    log(f"SUMMARY")
    log(f"  Invoices extracted: {invoice_count}")
    log(f"  Supporting docs:    {support_count}")
    log(f"  High confidence:    {high_conf}")
    log(f"  Low confidence:     {low_conf} (review recommended)")
    log(f"{'=' * 60}")

    return xlsx_path


# ===========================================================================
# GUI
# ===========================================================================
class Phase2GUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("PDF Invoice Processor — Phase 2: Vision Extraction")
        self.root.geometry("850x750")
        self.root.minsize(720, 600)

        self.manifest_path = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.ollama_url = tk.StringVar(value="http://192.168.1.101:11434")
        self.model_name = tk.StringVar(value="qwen3-vl:8b")
        self.include_ocr = tk.BooleanVar(value=True)
        self.processing = False
        self.cancel_event = threading.Event()

        self._build_ui()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main, text="PDF Invoice Processor — Phase 2",
                  font=("Segoe UI", 16, "bold")).pack(pady=(0, 2))
        ttk.Label(main,
                  text="Classify · Group · Extract — Vision model via Ollama (100% local)",
                  font=("Segoe UI", 10)).pack(pady=(0, 10))

        # --- Ollama settings ---
        ollama_frame = ttk.LabelFrame(main, text="Ollama Settings", padding=5)
        ollama_frame.pack(fill=tk.X, pady=(0, 5))

        row1 = ttk.Frame(ollama_frame)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="URL:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(row1, textvariable=self.ollama_url, width=30).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(row1, text="Model:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(row1, textvariable=self.model_name, width=20).pack(side=tk.LEFT, padx=(0, 10))
        self.ollama_status = ttk.Label(row1, text="")
        self.ollama_status.pack(side=tk.LEFT)
        ttk.Button(row1, text="Test", command=self._test_ollama).pack(side=tk.RIGHT)

        row2 = ttk.Frame(ollama_frame)
        row2.pack(fill=tk.X, pady=2)
        ttk.Checkbutton(row2, text="Include OCR text as supplementary context (recommended)",
                        variable=self.include_ocr).pack(anchor=tk.W)

        # --- Input ---
        input_frame = ttk.LabelFrame(main, text="Phase 1 Output (manifest.json)", padding=5)
        input_frame.pack(fill=tk.X, pady=(0, 5))
        input_row = ttk.Frame(input_frame)
        input_row.pack(fill=tk.X)
        ttk.Entry(input_row, textvariable=self.manifest_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(input_row, text="Browse…", command=self._browse_manifest).pack(side=tk.RIGHT)

        # --- Output ---
        output_frame = ttk.LabelFrame(main, text="Output Folder", padding=5)
        output_frame.pack(fill=tk.X, pady=(0, 5))
        output_row = ttk.Frame(output_frame)
        output_row.pack(fill=tk.X)
        ttk.Entry(output_row, textvariable=self.output_folder).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_row, text="Browse…", command=self._browse_output).pack(side=tk.RIGHT)

        # --- Progress ---
        prog_frame = ttk.LabelFrame(main, text="Progress", padding=5)
        prog_frame.pack(fill=tk.X, pady=(0, 5))
        self.progress_label = ttk.Label(prog_frame, text="Ready")
        self.progress_label.pack(anchor=tk.W)
        self.progress_bar = ttk.Progressbar(prog_frame, mode="determinate")
        self.progress_bar.pack(fill=tk.X, pady=(3, 0))

        # --- Buttons ---
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(5, 5))
        self.start_btn = ttk.Button(btn_frame, text="▶  Start Extraction",
                                    command=self._start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.cancel_btn = ttk.Button(btn_frame, text="■  Cancel",
                                     command=self._cancel, state=tk.DISABLED)
        self.cancel_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.open_btn = ttk.Button(btn_frame, text="Open Output",
                                   command=self._open_output, state=tk.DISABLED)
        self.open_btn.pack(side=tk.RIGHT)

        # --- Log ---
        log_frame = ttk.LabelFrame(main, text="Processing Log", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=14, wrap=tk.WORD,
                                                   font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _test_ollama(self):
        url = self.ollama_url.get().strip()
        model = self.model_name.get().strip()
        client = OllamaClient(base_url=url, model=model)

        if client.is_available():
            self.ollama_status.config(text="✓ Connected (vision)", foreground="green")
        else:
            models = client.list_models()
            if models:
                self.ollama_status.config(
                    text=f"⚠ Not found. Have: {', '.join(models[:5])}",
                    foreground="orange")
            else:
                self.ollama_status.config(text="✗ Cannot reach Ollama", foreground="red")

    def _browse_manifest(self):
        path = filedialog.askopenfilename(
            title="Select manifest.json from Phase 1",
            filetypes=[("JSON", "*.json"), ("All", "*.*")],
        )
        if path:
            self.manifest_path.set(path)
            if not self.output_folder.get():
                self.output_folder.set(
                    os.path.join(os.path.dirname(path), "extraction_output"))

    def _browse_output(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self.output_folder.set(folder)

    def _start_processing(self):
        manifest = self.manifest_path.get().strip()
        output = self.output_folder.get().strip()

        if not manifest or not os.path.isfile(manifest):
            messagebox.showerror("Error", "Select a valid manifest.json.")
            return
        if not output:
            messagebox.showerror("Error", "Select an output folder.")
            return

        self.log_text.delete("1.0", tk.END)
        self.progress_bar["value"] = 0
        self.cancel_event.clear()
        self.processing = True
        self.start_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL)
        self.open_btn.config(state=tk.DISABLED)

        thread = threading.Thread(target=self._run_thread,
                                  args=(manifest, output), daemon=True)
        thread.start()

    def _run_thread(self, manifest_path, output_dir):
        try:
            result = run_phase2(
                manifest_path=manifest_path,
                output_dir=output_dir,
                ollama_url=self.ollama_url.get().strip(),
                model=self.model_name.get().strip(),
                include_ocr_text=self.include_ocr.get(),
                log_callback=self._append_log,
                progress_callback=self._update_progress,
                cancel_event=self.cancel_event,
            )
            self.root.after(0, self._done, result)
        except Exception as e:
            self.root.after(0, self._error, str(e))

    def _cancel(self):
        self.cancel_event.set()
        self._append_log("Cancelling...")

    def _done(self, result_path):
        self.processing = False
        self.start_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED)
        self.open_btn.config(state=tk.NORMAL)
        self.progress_bar["value"] = 100
        if result_path:
            self.progress_label.config(text=f"Complete → {os.path.basename(result_path)}")
        else:
            self.progress_label.config(text="Completed with errors — check log")

    def _error(self, msg):
        self.processing = False
        self.start_btn.config(state=tk.NORMAL)
        self.cancel_btn.config(state=tk.DISABLED)
        messagebox.showerror("Error", msg)

    def _update_progress(self, current, total, msg):
        def _u():
            pct = (current / total * 100) if total > 0 else 0
            self.progress_bar["value"] = pct
            self.progress_label.config(text=f"[{current}/{total}] {msg}")
        self.root.after(0, _u)

    def _append_log(self, message):
        def _a():
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
        self.root.after(0, _a)

    def _open_output(self):
        folder = self.output_folder.get()
        if folder and os.path.isdir(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", folder])
            else:
                subprocess.Popen(["xdg-open", folder])


def main():
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
    root = tk.Tk()
    try:
        style = ttk.Style()
        for t in ("clam", "alt", "vista"):
            if t in style.theme_names():
                style.theme_use(t)
                break
    except Exception:
        pass
    Phase2GUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
