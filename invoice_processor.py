"""
PDF Invoice Processor — Unified Pipeline

Combines document preparation (splitting, orientation, OCR) and
AI-powered extraction (classify, group, extract) into a single
tabbed application.

Tab 1 — Document Preparation:
  Reads PDFs, splits multi-page files into single pages, detects and
  corrects orientation, applies OCR where needed.

Tab 2 — Vision Extraction:
  Uses Ollama with a vision model to classify pages, group them into
  invoices, and extract structured fields to CSV/XLSX.

All processing is done locally — no data leaves the machine.

Requirements:
    System: tesseract-ocr, poppler-utils, ghostscript
    Python: pypdf, pytesseract, pdf2image, Pillow, ocrmypdf, pdfplumber,
            tqdm, deskew, numpy, openpyxl
    Ollama: running locally with a vision model (e.g. qwen3-vl:8b)
"""

import os
import sys
import csv
import json
import shutil
import logging
import subprocess
import threading
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict, field
from typing import Optional

# ---------------------------------------------------------------------------
# PDF / Image / OCR imports
# ---------------------------------------------------------------------------
try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    sys.exit("Missing 'pypdf'. Run: pip install pypdf")

try:
    from PIL import Image
except ImportError:
    sys.exit("Missing 'Pillow'. Run: pip install Pillow")

try:
    import pytesseract
except ImportError:
    sys.exit("Missing 'pytesseract'. Run: pip install pytesseract")

try:
    from pdf2image import convert_from_path
except ImportError:
    sys.exit("Missing 'pdf2image'. Run: pip install pdf2image")

try:
    import pdfplumber
except ImportError:
    sys.exit("Missing 'pdfplumber'. Run: pip install pdfplumber")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing 'openpyxl'. Run: pip install openpyxl")

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

from ollama_client import OllamaClient, OllamaError
from prompts import (
    SYSTEM_PROMPT,
    CLASSIFY_PAGE_PROMPT,
    CLASSIFY_PAGE_WITH_TEXT_PROMPT,
    GROUP_PAGES_PROMPT,
    EXTRACT_FIELDS_PROMPT,
    EXTRACT_FIELDS_WITH_TEXT_PROMPT,
    EXTRACT_KEY_FIELDS_PROMPT,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
TEXT_THRESHOLD = 40
OCR_DPI = 300
ORIENTATION_CONFIDENCE = 2.0
LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
MAX_TEXT_LENGTH = 4000
MAX_IMAGES_PER_EXTRACTION = 4
LOW_CONFIDENCE_RETRY_THRESHOLD = 0.6


# ===========================================================================
# Data classes
# ===========================================================================
@dataclass
class PageRecord:
    """Metadata for a single processed page (Phase 1 output)."""
    page_id: str
    source_pdf: str
    source_pdf_basename: str
    original_page: int
    total_source_pages: int
    orientation_detected: int
    orientation_correction: int
    orientation_confidence: float
    had_text: bool
    ocr_applied: bool
    text_length: int
    output_pdf: str
    output_image: str
    processing_notes: list = field(default_factory=list)


@dataclass
class PageClassification:
    page_id: str
    source_pdf: str
    original_page: int
    page_type: str
    confidence: float
    invoice_number: Optional[str]
    page_indicator: Optional[str]
    reasoning: str
    image_path: str


@dataclass
class InvoiceGroup:
    group_id: int
    group_type: str
    invoice_number: Optional[str]
    page_ids: list
    pages: list


@dataclass
class InvoiceRecord:
    group_id: int
    page_ids: list
    source_files: list
    invoice_number: Optional[str]
    invoice_date: Optional[str]
    due_date: Optional[str]
    supplier_name: Optional[str]
    supplier_address: Optional[str]
    customer_name: Optional[str]
    customer_address: Optional[str]
    purchase_order_number: Optional[str]
    currency: Optional[str]
    line_items_summary: str
    subtotal: Optional[float]
    shipping: Optional[float]
    gst: Optional[float]
    pst: Optional[float]
    hst: Optional[float]
    other_taxes: Optional[float]
    total: Optional[float]
    notes: Optional[str]
    extraction_confidence: float
    fields_uncertain: list
    num_pages: int
    group_type: str


# ===========================================================================
# Shared helpers
# ===========================================================================
def extract_page_text(pdf_path: str, page_num: int = 0) -> str:
    """Extract embedded text from a single page of a PDF using pdfplumber."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_num < len(pdf.pages):
                text = pdf.pages[page_num].extract_text() or ""
                return text.strip()
    except Exception:
        pass
    return ""


def truncate_text(text: str, max_len: int = MAX_TEXT_LENGTH) -> str:
    if len(text) <= max_len:
        return text
    half = max_len // 2
    return text[:half] + "\n\n[... truncated ...]\n\n" + text[-half:]


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _first_non_empty(*vals):
    for val in vals:
        if val is None:
            continue
        if isinstance(val, str):
            if val.strip():
                return val.strip()
        else:
            return val
    return None


def check_system_dependencies() -> list[str]:
    """Return a list of missing system tools."""
    missing = []
    gs_candidates = ("gswin64c", "gswin64", "gs") if sys.platform == "win32" else ("gs",)
    for tool in ("tesseract", "pdftoppm"):
        if shutil.which(tool) is None:
            missing.append(tool)
    if not any(shutil.which(g) for g in gs_candidates):
        missing.append("gs")
    return missing


def open_folder(folder: str):
    """Open a folder in the system file manager."""
    if folder and os.path.isdir(folder):
        if sys.platform == "win32":
            os.startfile(folder)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", folder])
        else:
            subprocess.Popen(["xdg-open", folder])


# ===========================================================================
# Phase 1: Document Preparation — pipeline functions
# ===========================================================================
def render_page_to_image(pdf_path: str, page_num: int = 0, dpi: int = OCR_DPI) -> Image.Image:
    images = convert_from_path(pdf_path, first_page=page_num + 1, last_page=page_num + 1, dpi=dpi)
    return images[0]


def detect_orientation(image: Image.Image) -> tuple[int, float]:
    try:
        osd = pytesseract.image_to_osd(image, output_type=pytesseract.Output.DICT)
        rotate = osd.get("rotate", 0)
        confidence = float(osd.get("orientation_conf", 0.0))
        return rotate, confidence
    except pytesseract.TesseractError:
        return 0, 0.0


def rotate_image(image: Image.Image, degrees: int) -> Image.Image:
    if degrees == 0:
        return image
    return image.rotate(-degrees, expand=True)


def deskew_image_pil(image: Image.Image) -> tuple[Image.Image, float]:
    try:
        import numpy as np
        from deskew import determine_skew
        gray = np.array(image.convert("L"))
        angle = determine_skew(gray)
        if angle is not None and abs(angle) > 0.3:
            corrected = image.rotate(angle, expand=True, fillcolor="white")
            return corrected, round(angle, 2)
    except ImportError:
        pass
    except Exception as e:
        logging.debug(f"Deskew failed: {e}")
    return image, 0.0


def ocr_image_to_pdf_bytes(image: Image.Image) -> bytes:
    return pytesseract.image_to_pdf_or_hocr(image, extension="pdf")


def create_single_page_pdf(reader: PdfReader, page_index: int, output_path: str):
    writer = PdfWriter()
    writer.add_page(reader.pages[page_index])
    with open(output_path, "wb") as f:
        writer.write(f)


def apply_rotation_to_pdf(input_pdf: str, degrees: int, output_pdf: str):
    reader = PdfReader(input_pdf)
    writer = PdfWriter()
    page = reader.pages[0]
    page.rotate(degrees)
    writer.add_page(page)
    with open(output_pdf, "wb") as f:
        writer.write(f)


def apply_ocr_to_pdf(input_pdf: str, output_pdf: str, force_ocr: bool = True) -> bool:
    try:
        import ocrmypdf
        ocrmypdf.ocr(input_pdf, output_pdf, language="eng", force_ocr=force_ocr, optimize=1)
        return True
    except ImportError:
        return False
    except Exception as e:
        logging.warning(f"ocrmypdf failed: {e}, falling back to manual OCR")
        return False


def process_page(
    source_pdf_path: str,
    page_index: int,
    total_pages: int,
    output_dir: str,
    images_dir: str,
    page_counter: int,
) -> PageRecord:
    source_basename = os.path.basename(source_pdf_path)
    stem = Path(source_pdf_path).stem
    page_id = f"{stem}_page_{page_index + 1:04d}"

    notes = []
    temp_files = []

    try:
        reader = PdfReader(source_pdf_path)
        temp_single = os.path.join(output_dir, f"_temp_{page_id}.pdf")
        create_single_page_pdf(reader, page_index, temp_single)
        temp_files.append(temp_single)

        embedded_text = extract_page_text(temp_single, 0)
        has_text = len(embedded_text) >= TEXT_THRESHOLD

        if has_text:
            notes.append(f"Embedded text found ({len(embedded_text)} chars)")
        else:
            notes.append(f"No usable embedded text ({len(embedded_text)} chars)")

        page_image = render_page_to_image(temp_single, 0, dpi=OCR_DPI)

        rotation_needed, osd_confidence = detect_orientation(page_image)
        notes.append(f"OSD: rotate={rotation_needed}°, conf={osd_confidence:.1f}")

        actual_rotation = 0
        if osd_confidence >= ORIENTATION_CONFIDENCE and rotation_needed != 0:
            actual_rotation = rotation_needed
            notes.append(f"Applying {actual_rotation}° rotation")
        elif rotation_needed != 0:
            notes.append(f"Low confidence ({osd_confidence:.1f}), skipping rotation")

        corrected_image = rotate_image(page_image, actual_rotation)

        corrected_image, deskew_angle = deskew_image_pil(corrected_image)
        if deskew_angle:
            notes.append(f"Image deskew: {deskew_angle:.2f}°")

        image_path = os.path.join(images_dir, f"{page_id}.png")
        corrected_image.save(image_path, "PNG")

        output_pdf_path = os.path.join(output_dir, f"{page_id}.pdf")

        if deskew_angle:
            pdf_bytes = ocr_image_to_pdf_bytes(corrected_image)
            with open(output_pdf_path, "wb") as f:
                f.write(pdf_bytes)
            notes.append("PDF rebuilt from deskewed image via Tesseract")
            ocr_applied = True

        elif has_text and actual_rotation == 0:
            shutil.copy2(temp_single, output_pdf_path)
            ocr_applied = False
            notes.append("Kept original PDF (no corrections needed)")

        elif has_text and actual_rotation != 0:
            apply_rotation_to_pdf(temp_single, actual_rotation, output_pdf_path)
            ocr_applied = False
            notes.append("Rotated existing PDF")

        else:
            if actual_rotation != 0:
                rotated_temp = os.path.join(output_dir, f"_temp_rot_{page_id}.pdf")
                apply_rotation_to_pdf(temp_single, actual_rotation, rotated_temp)
                temp_files.append(rotated_temp)
                ocr_input = rotated_temp
            else:
                ocr_input = temp_single

            ocr_success = apply_ocr_to_pdf(ocr_input, output_pdf_path, force_ocr=True)
            if not ocr_success:
                pdf_bytes = ocr_image_to_pdf_bytes(corrected_image)
                with open(output_pdf_path, "wb") as f:
                    f.write(pdf_bytes)
                notes.append("OCR via Tesseract direct (ocrmypdf unavailable)")
            else:
                notes.append("OCR via ocrmypdf")
            ocr_applied = True

        final_text = extract_page_text(output_pdf_path, 0)

        return PageRecord(
            page_id=page_id,
            source_pdf=source_pdf_path,
            source_pdf_basename=source_basename,
            original_page=page_index + 1,
            total_source_pages=total_pages,
            orientation_detected=rotation_needed,
            orientation_correction=actual_rotation,
            orientation_confidence=osd_confidence,
            had_text=has_text,
            ocr_applied=ocr_applied,
            text_length=len(final_text),
            output_pdf=output_pdf_path,
            output_image=image_path,
            processing_notes=notes,
        )

    finally:
        for tf in temp_files:
            try:
                if os.path.exists(tf):
                    os.remove(tf)
            except OSError:
                pass


def discover_pdfs(folder: str) -> list[str]:
    pdfs = []
    for entry in sorted(os.listdir(folder)):
        if entry.lower().endswith(".pdf"):
            pdfs.append(os.path.join(folder, entry))
    return pdfs


def count_total_pages(pdf_paths: list[str]) -> int:
    total = 0
    for p in pdf_paths:
        try:
            reader = PdfReader(p)
            total += len(reader.pages)
        except Exception:
            pass
    return total


def run_phase1(
    input_folder: str,
    output_folder: str,
    recursive: bool = False,
    progress_callback=None,
    log_callback=None,
    cancel_event: threading.Event = None,
) -> list[PageRecord]:
    """Main Phase 1 pipeline."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    def progress(current, total, msg=""):
        if progress_callback:
            progress_callback(current, total, msg)

    if recursive:
        pdf_paths = []
        for root, _, files in os.walk(input_folder):
            for f in sorted(files):
                if f.lower().endswith(".pdf"):
                    pdf_paths.append(os.path.join(root, f))
    else:
        pdf_paths = discover_pdfs(input_folder)

    if not pdf_paths:
        log("No PDF files found in the selected folder.")
        return []

    log(f"Found {len(pdf_paths)} PDF file(s)")

    total_pages = count_total_pages(pdf_paths)
    log(f"Total pages to process: {total_pages}")

    pages_dir = os.path.join(output_folder, "pages")
    images_dir = os.path.join(output_folder, "images")
    os.makedirs(pages_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)

    records = []
    page_counter = 0

    for pdf_path in pdf_paths:
        if cancel_event and cancel_event.is_set():
            log("Processing cancelled by user.")
            break

        try:
            reader = PdfReader(pdf_path)
            num_pages = len(reader.pages)
        except Exception as e:
            log(f"ERROR: Cannot read '{os.path.basename(pdf_path)}': {e}")
            continue

        log(f"Processing: {os.path.basename(pdf_path)} ({num_pages} pages)")

        for page_idx in range(num_pages):
            if cancel_event and cancel_event.is_set():
                break

            page_counter += 1
            progress(page_counter, total_pages,
                     f"{os.path.basename(pdf_path)} p.{page_idx + 1}/{num_pages}")

            try:
                record = process_page(
                    source_pdf_path=pdf_path,
                    page_index=page_idx,
                    total_pages=num_pages,
                    output_dir=pages_dir,
                    images_dir=images_dir,
                    page_counter=page_counter,
                )
                records.append(record)

                status = "OCR" if record.ocr_applied else "TEXT"
                rot = f" rot:{record.orientation_correction}°" if record.orientation_correction else ""
                log(f"  Page {page_idx + 1}: [{status}]{rot} → {record.page_id}.pdf "
                    f"({record.text_length} chars)")

            except Exception as e:
                log(f"  ERROR on page {page_idx + 1}: {e}")
                logging.exception(f"Failed processing {pdf_path} page {page_idx}")

    manifest_path = os.path.join(output_folder, "manifest.json")
    manifest_data = {
        "created": datetime.now().isoformat(),
        "input_folder": input_folder,
        "output_folder": output_folder,
        "total_source_files": len(pdf_paths),
        "total_pages_processed": len(records),
        "pages": [asdict(r) for r in records],
    }
    with open(manifest_path, "w") as f:
        json.dump(manifest_data, f, indent=2)

    log(f"\nDone! Processed {len(records)} pages from {len(pdf_paths)} file(s)")
    log(f"Output: {output_folder}")
    log(f"Manifest: {manifest_path}")

    return records


# ===========================================================================
# Phase 2: Vision Extraction — pipeline functions
# ===========================================================================
def _build_extraction_prompt(base_prompt: str, ordered_page_ids: list[str]) -> str:
    if len(ordered_page_ids) <= 1:
        return base_prompt
    page_order_lines = "\n".join(
        f"- Image {idx + 1}: {page_id}" for idx, page_id in enumerate(ordered_page_ids)
    )
    return (
        f"{base_prompt}\n\n"
        "Important multi-page guidance:\n"
        "The images are consecutive invoice pages in this exact order:\n"
        f"{page_order_lines}\n"
        "- Use page 1 for header fields (supplier/customer, invoice number/date).\n"
        "- Use later pages for continued line items, taxes, and totals.\n"
        "- Return a single consolidated JSON for the whole invoice."
    )


def _merge_two_page_rescue(primary: dict, page1: dict, page2: dict) -> dict:
    merged = dict(primary)

    text_fields = [
        "invoice_number", "invoice_date", "due_date",
        "supplier_name", "supplier_address",
        "customer_name", "customer_address",
        "purchase_order_number", "currency", "notes",
    ]
    for f in text_fields:
        merged[f] = _first_non_empty(primary.get(f), page1.get(f), page2.get(f))

    amount_fields = ["subtotal", "shipping", "gst", "pst", "hst", "other_taxes", "total"]
    for f in amount_fields:
        merged[f] = _first_non_empty(
            _to_float(primary.get(f)),
            _to_float(page2.get(f)),
            _to_float(page1.get(f)),
        )

    if not isinstance(merged.get("line_items"), list):
        merged["line_items"] = []

    uncertain = merged.get("fields_uncertain", [])
    if not isinstance(uncertain, list):
        uncertain = []
    resolved = set()
    for f in text_fields + amount_fields:
        val = merged.get(f)
        if val is not None and (not isinstance(val, str) or val.strip()):
            resolved.add(f)
    merged["fields_uncertain"] = [f for f in uncertain if f not in resolved]

    primary_conf = float(primary.get("extraction_confidence", 0.0) or 0.0)
    page1_conf = float(page1.get("extraction_confidence", 0.0) or 0.0)
    page2_conf = float(page2.get("extraction_confidence", 0.0) or 0.0)
    merged["extraction_confidence"] = max(
        primary_conf,
        min(0.85, (primary_conf * 0.5) + (page1_conf * 0.25) + (page2_conf * 0.25) + 0.1),
    )
    return merged


def classify_pages(
    manifest: dict,
    client: OllamaClient,
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> list[PageClassification]:

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    pages = manifest.get("pages", [])
    total = len(pages)
    classifications = []

    log(f"Classifying {total} pages using vision model...")

    for i, page_data in enumerate(pages):
        if cancel_event and cancel_event.is_set():
            break

        page_id = page_data["page_id"]
        image_path = page_data.get("output_image", "")
        pdf_path = page_data.get("output_pdf", "")

        if progress_callback:
            progress_callback(i + 1, total, f"Classifying {page_id}")

        if not image_path or not os.path.exists(image_path):
            log(f"  {page_id}: No image found — classifying as unknown")
            classifications.append(PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type="unknown", confidence=0.0,
                invoice_number=None, page_indicator=None,
                reasoning="No image available", image_path="",
            ))
            continue

        if include_ocr_text and pdf_path and os.path.exists(pdf_path):
            ocr_text = extract_page_text(pdf_path)
            if len(ocr_text) > 20:
                prompt = CLASSIFY_PAGE_WITH_TEXT_PROMPT.format(
                    page_text=truncate_text(ocr_text)
                )
            else:
                prompt = CLASSIFY_PAGE_PROMPT
        else:
            prompt = CLASSIFY_PAGE_PROMPT

        try:
            result = client.generate_json_with_image(
                prompt=prompt,
                image_paths=[image_path],
                system=SYSTEM_PROMPT,
                timeout=120,
            )

            cls = PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type=result.get("page_type", "unknown"),
                confidence=float(result.get("confidence", 0.0)),
                invoice_number=result.get("invoice_number"),
                page_indicator=result.get("page_indicator"),
                reasoning=result.get("reasoning", ""),
                image_path=image_path,
            )
            classifications.append(cls)

            inv = f" INV#{cls.invoice_number}" if cls.invoice_number else ""
            log(f"  {page_id}: {cls.page_type} (conf:{cls.confidence:.2f}){inv}")

        except OllamaError as e:
            log(f"  {page_id}: ERROR — {e}")
            classifications.append(PageClassification(
                page_id=page_id,
                source_pdf=page_data.get("source_pdf_basename", ""),
                original_page=page_data.get("original_page", 0),
                page_type="unknown", confidence=0.0,
                invoice_number=None, page_indicator=None,
                reasoning=f"VLM error: {e}", image_path=image_path,
            ))

    return classifications


def group_pages(
    classifications: list[PageClassification],
    client: OllamaClient,
    log_callback=None,
) -> list[InvoiceGroup]:

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    pages_summary = []
    for cls in classifications:
        pages_summary.append({
            "page_id": cls.page_id,
            "page_type": cls.page_type,
            "invoice_number": cls.invoice_number,
            "page_indicator": cls.page_indicator,
            "source_pdf": cls.source_pdf,
            "original_page": cls.original_page,
        })

    chunk_size = 15
    all_groups = []
    group_id_counter = 1

    for start in range(0, len(pages_summary), chunk_size):
        chunk = pages_summary[start:start + chunk_size]
        log(f"Grouping pages {start + 1}-{start + len(chunk)} of {len(pages_summary)}...")

        prompt = GROUP_PAGES_PROMPT.format(pages_json=json.dumps(chunk, indent=2))

        success = False
        for attempt in range(2):
            try:
                result = client.generate_json(prompt, system=SYSTEM_PROMPT, timeout=600)
                raw_groups = result.get("groups", [])

                for g in raw_groups:
                    page_ids = g.get("page_ids", [])
                    group_pages_list = [c for c in classifications if c.page_id in page_ids]
                    all_groups.append(InvoiceGroup(
                        group_id=group_id_counter,
                        group_type=g.get("group_type", "invoice"),
                        invoice_number=g.get("invoice_number"),
                        page_ids=page_ids,
                        pages=group_pages_list,
                    ))
                    group_id_counter += 1

                success = True
                break

            except OllamaError as e:
                if attempt == 0:
                    log(f"  Grouping attempt {attempt + 1} failed ({e}), retrying...")
                else:
                    log(f"  Grouping error: {e}")
                    log("  Falling back to sequential grouping...")

        if not success:
            all_groups.extend(
                _fallback_grouping(classifications[start:start + chunk_size], group_id_counter)
            )
            group_id_counter = max((g.group_id for g in all_groups), default=0) + 1

    grouped_ids = set()
    for g in all_groups:
        grouped_ids.update(g.page_ids)

    ungrouped = [c for c in classifications if c.page_id not in grouped_ids]
    if ungrouped:
        log(f"  {len(ungrouped)} pages not grouped — adding individually")
        for cls in ungrouped:
            all_groups.append(InvoiceGroup(
                group_id=group_id_counter,
                group_type="invoice" if "invoice" in cls.page_type else "supporting_documents",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            ))
            group_id_counter += 1

    log(f"Created {len(all_groups)} groups")
    for g in all_groups:
        inv = f" INV#{g.invoice_number}" if g.invoice_number else ""
        log(f"  Group {g.group_id}: {g.group_type}{inv} — {len(g.page_ids)} page(s)")

    return all_groups


def _fallback_grouping(classifications, start_id):
    groups = []
    current = None
    gid = start_id

    for cls in classifications:
        if cls.page_type == "invoice_first_page":
            if current:
                groups.append(current)
            current = InvoiceGroup(
                group_id=gid, group_type="invoice",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            )
            gid += 1
        elif cls.page_type == "invoice_continuation" and current:
            current.page_ids.append(cls.page_id)
            current.pages.append(cls)
        else:
            if current:
                groups.append(current)
                current = None
            groups.append(InvoiceGroup(
                group_id=gid,
                group_type="supporting_documents" if cls.page_type == "supporting_document" else "unknown",
                invoice_number=cls.invoice_number,
                page_ids=[cls.page_id], pages=[cls],
            ))
            gid += 1

    if current:
        groups.append(current)
    return groups


def extract_invoice_fields(
    groups: list[InvoiceGroup],
    manifest: dict,
    client: OllamaClient,
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> list[InvoiceRecord]:

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    page_image_map = {}
    page_pdf_map = {}
    for p in manifest.get("pages", []):
        page_image_map[p["page_id"]] = p.get("output_image", "")
        page_pdf_map[p["page_id"]] = p.get("output_pdf", "")
    page_order_map = {p["page_id"]: i for i, p in enumerate(manifest.get("pages", []))}

    records = []
    invoice_groups = [g for g in groups if g.group_type == "invoice"]
    total = len(invoice_groups)

    log(f"\nExtracting fields from {total} invoice group(s) using vision model...")

    for i, group in enumerate(invoice_groups):
        if cancel_event and cancel_event.is_set():
            break

        if progress_callback:
            progress_callback(i + 1, total, f"Extracting group {group.group_id}")

        inv_label = f"Group {group.group_id}"
        if group.invoice_number:
            inv_label += f" (INV#{group.invoice_number})"

        image_paths = []
        source_files = set()
        ocr_texts = []

        ordered_page_ids = sorted(
            group.page_ids,
            key=lambda pid: page_order_map.get(pid, 10**9),
        )

        for page_id in ordered_page_ids:
            img = page_image_map.get(page_id, "")
            if img and os.path.exists(img):
                image_paths.append(img)
            pdf = page_pdf_map.get(page_id, "")
            if include_ocr_text and pdf and os.path.exists(pdf):
                text = extract_page_text(pdf)
                if text:
                    ocr_texts.append(f"--- {page_id} ---\n{text}")
            for p in group.pages:
                if p.page_id == page_id:
                    source_files.add(p.source_pdf)

        if not image_paths:
            log(f"  {inv_label}: No images — skipping")
            continue

        if len(image_paths) > MAX_IMAGES_PER_EXTRACTION:
            log(f"  {inv_label}: {len(image_paths)} pages, sending first {MAX_IMAGES_PER_EXTRACTION}")
            image_paths = image_paths[:MAX_IMAGES_PER_EXTRACTION]

        combined_ocr = "\n".join(ocr_texts)
        if combined_ocr and len(combined_ocr) > 30:
            base_prompt = EXTRACT_FIELDS_WITH_TEXT_PROMPT.format(
                invoice_text=truncate_text(combined_ocr, 6000)
            )
        else:
            base_prompt = EXTRACT_FIELDS_PROMPT
        prompt = _build_extraction_prompt(base_prompt, ordered_page_ids)

        try:
            result = client.generate_json_with_image(
                prompt=prompt,
                image_paths=image_paths,
                system=SYSTEM_PROMPT,
                timeout=240,
            )
        except OllamaError as e:
            log(f"  {inv_label}: Full extraction failed ({e}), trying simplified...")
            try:
                result = client.generate_json_with_image(
                    prompt=EXTRACT_KEY_FIELDS_PROMPT,
                    image_paths=image_paths[:1],
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                result.setdefault("extraction_confidence", 0.3)
                result.setdefault("fields_uncertain", ["most fields"])
            except OllamaError as e2:
                log(f"  {inv_label}: All extraction failed — {e2}")
                continue

        result_conf = float(result.get("extraction_confidence", 0.0) or 0.0)
        if len(image_paths) == 2 and result_conf < LOW_CONFIDENCE_RETRY_THRESHOLD:
            log(f"  {inv_label}: Low confidence ({result_conf:.2f}) - running page-wise rescue")
            try:
                page1_result = client.generate_json_with_image(
                    prompt=EXTRACT_KEY_FIELDS_PROMPT,
                    image_paths=[image_paths[0]],
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                page2_result = client.generate_json_with_image(
                    prompt=(
                        EXTRACT_KEY_FIELDS_PROMPT
                        + "\n\nPrioritize taxes and totals if they appear on this page."
                    ),
                    image_paths=[image_paths[1]],
                    system=SYSTEM_PROMPT,
                    timeout=180,
                )
                result = _merge_two_page_rescue(result, page1_result, page2_result)
            except OllamaError as rescue_error:
                log(f"  {inv_label}: Rescue pass failed - {rescue_error}")

        line_items = result.get("line_items", [])
        if isinstance(line_items, list) and line_items:
            items_summary = "; ".join(
                item.get("description", "?") for item in line_items[:10]
            )
            if len(line_items) > 10:
                items_summary += f" (+{len(line_items) - 10} more)"
        else:
            items_summary = ""

        record = InvoiceRecord(
            group_id=group.group_id,
            page_ids=ordered_page_ids,
            source_files=sorted(source_files),
            invoice_number=result.get("invoice_number") or group.invoice_number,
            invoice_date=result.get("invoice_date"),
            due_date=result.get("due_date"),
            supplier_name=result.get("supplier_name"),
            supplier_address=result.get("supplier_address"),
            customer_name=result.get("customer_name"),
            customer_address=result.get("customer_address"),
            purchase_order_number=result.get("purchase_order_number"),
            currency=result.get("currency"),
            line_items_summary=items_summary,
            subtotal=_to_float(result.get("subtotal")),
            shipping=_to_float(result.get("shipping")),
            gst=_to_float(result.get("gst")),
            pst=_to_float(result.get("pst")),
            hst=_to_float(result.get("hst")),
            other_taxes=_to_float(result.get("other_taxes")),
            total=_to_float(result.get("total")),
            notes=result.get("notes"),
            extraction_confidence=float(result.get("extraction_confidence", 0.0)),
            fields_uncertain=result.get("fields_uncertain", []),
            num_pages=len(ordered_page_ids),
            group_type=group.group_type,
        )
        records.append(record)

        conf_pct = f"{record.extraction_confidence * 100:.0f}%"
        log(f"  {inv_label}: {record.supplier_name or '?'} → ${record.total or '?'} "
            f"(conf:{conf_pct})")

    for group in groups:
        if group.group_type != "invoice":
            records.append(InvoiceRecord(
                group_id=group.group_id,
                page_ids=group.page_ids,
                source_files=sorted(set(p.source_pdf for p in group.pages)),
                invoice_number=None,
                invoice_date=None, due_date=None,
                supplier_name=None, supplier_address=None,
                customer_name=None, customer_address=None,
                purchase_order_number=None, currency=None,
                line_items_summary="",
                subtotal=None, shipping=None,
                gst=None, pst=None, hst=None, other_taxes=None, total=None,
                notes=f"Supporting document ({len(group.page_ids)} page(s))",
                extraction_confidence=0.0, fields_uncertain=[],
                num_pages=len(group.page_ids), group_type=group.group_type,
            ))

    return records


# ===========================================================================
# Output writers
# ===========================================================================
CSV_COLUMNS = [
    "group_id", "group_type", "invoice_number", "invoice_date", "due_date",
    "supplier_name", "supplier_address", "customer_name", "customer_address",
    "purchase_order_number", "currency", "line_items_summary",
    "subtotal", "shipping", "gst", "pst", "hst", "other_taxes", "total",
    "notes", "extraction_confidence", "fields_uncertain",
    "num_pages", "page_ids", "source_files",
]


def write_csv(records: list[InvoiceRecord], output_path: str):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for rec in records:
            row = asdict(rec)
            row["page_ids"] = "; ".join(row["page_ids"])
            row["source_files"] = "; ".join(row["source_files"])
            row["fields_uncertain"] = "; ".join(row["fields_uncertain"])
            writer.writerow({k: row.get(k) for k in CSV_COLUMNS})


def write_xlsx(records: list[InvoiceRecord], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    money_fmt = '#,##0.00'
    pct_fmt = '0%'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    col_config = [
        ("Group", 8, None),
        ("Type", 14, None),
        ("Invoice #", 16, None),
        ("Invoice Date", 14, None),
        ("Due Date", 14, None),
        ("Supplier", 25, None),
        ("Supplier Address", 30, None),
        ("Customer", 25, None),
        ("Customer Address", 30, None),
        ("PO #", 14, None),
        ("Currency", 10, None),
        ("Line Items", 45, None),
        ("Subtotal", 14, money_fmt),
        ("Shipping", 12, money_fmt),
        ("GST", 12, money_fmt),
        ("PST", 12, money_fmt),
        ("HST", 12, money_fmt),
        ("Other Tax", 12, money_fmt),
        ("Total", 14, money_fmt),
        ("Notes", 35, None),
        ("Confidence", 12, pct_fmt),
        ("Uncertain Fields", 25, None),
        ("Pages", 8, None),
        ("Page IDs", 35, None),
        ("Source Files", 30, None),
    ]

    for col_idx, (label, width, _) in enumerate(col_config, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "A2"

    invoice_fill = PatternFill("solid", fgColor="FFFFFF")
    support_fill = PatternFill("solid", fgColor="F2F2F2")
    low_conf_fill = PatternFill("solid", fgColor="FFF2CC")

    for row_idx, rec in enumerate(records, start=2):
        row_data = [
            rec.group_id, rec.group_type, rec.invoice_number,
            rec.invoice_date, rec.due_date,
            rec.supplier_name, rec.supplier_address,
            rec.customer_name, rec.customer_address,
            rec.purchase_order_number, rec.currency,
            rec.line_items_summary,
            rec.subtotal, rec.shipping,
            rec.gst, rec.pst, rec.hst, rec.other_taxes, rec.total,
            rec.notes, rec.extraction_confidence,
            "; ".join(rec.fields_uncertain) if rec.fields_uncertain else "",
            rec.num_pages,
            "; ".join(rec.page_ids),
            "; ".join(rec.source_files),
        ]

        bg = support_fill if rec.group_type != "invoice" else invoice_fill
        if rec.group_type == "invoice" and rec.extraction_confidence < 0.5:
            bg = low_conf_fill

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.fill = bg
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            _, _, num_fmt = col_config[col_idx - 1]
            if num_fmt and value is not None:
                cell.number_format = num_fmt

    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(col_config)).column_letter}{len(records) + 1}"
    wb.save(output_path)


def run_phase2(
    manifest_path: str,
    output_dir: str,
    ollama_url: str = "http://192.168.1.101:11434",
    model: str = "qwen3-vl:8b",
    include_ocr_text: bool = True,
    log_callback=None,
    progress_callback=None,
    cancel_event: threading.Event = None,
) -> Optional[str]:
    """Run the full Phase 2 pipeline. Returns path to output XLSX or None."""

    def log(msg):
        logging.info(msg)
        if log_callback:
            log_callback(msg)

    log(f"Loading manifest: {manifest_path}")
    with open(manifest_path, "r") as f:
        manifest = json.load(f)

    total_pages = len(manifest.get("pages", []))
    log(f"Manifest has {total_pages} pages from {manifest.get('total_source_files', '?')} file(s)")

    images_found = sum(
        1 for p in manifest.get("pages", [])
        if os.path.exists(p.get("output_image", ""))
    )
    log(f"Page images found: {images_found}/{total_pages}")

    if images_found == 0:
        log("ERROR: No page images found. Did Phase 1 run successfully?")
        return None

    client = OllamaClient(base_url=ollama_url, model=model)
    log(f"Connecting to Ollama at {ollama_url} with model {model}...")

    if not client.is_available():
        available = client.list_models()
        if available:
            log(f"ERROR: Model '{model}' not found. Available: {', '.join(available)}")
        else:
            log(f"ERROR: Cannot reach Ollama at {ollama_url}. Is it running?")
            log("  Start with: ollama serve")
            log(f"  Pull model: ollama pull {model}")
        return None

    log(f"Ollama connected — vision model '{model}' ready")
    ocr_mode = "with OCR text supplement" if include_ocr_text else "vision only"
    log(f"Mode: {ocr_mode}")

    os.makedirs(output_dir, exist_ok=True)

    # Step 1: Classify
    log(f"\n{'=' * 60}")
    log("STEP 1: Page Classification (vision)")
    log(f"{'=' * 60}")
    classifications = classify_pages(
        manifest, client,
        include_ocr_text=include_ocr_text,
        log_callback=log_callback,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
    )
    if cancel_event and cancel_event.is_set():
        return None

    cls_path = os.path.join(output_dir, "classifications.json")
    with open(cls_path, "w") as f:
        json.dump([asdict(c) for c in classifications], f, indent=2)
    log(f"Saved → {cls_path}")

    # Step 2: Group
    log(f"\n{'=' * 60}")
    log("STEP 2: Page Grouping")
    log(f"{'=' * 60}")
    groups = group_pages(classifications, client, log_callback=log_callback)
    if cancel_event and cancel_event.is_set():
        return None

    groups_path = os.path.join(output_dir, "groups.json")
    groups_data = []
    for g in groups:
        gd = asdict(g)
        gd.pop("pages", None)
        groups_data.append(gd)
    with open(groups_path, "w") as f:
        json.dump(groups_data, f, indent=2)
    log(f"Saved → {groups_path}")

    # Step 3: Extract
    log(f"\n{'=' * 60}")
    log("STEP 3: Field Extraction (vision)")
    log(f"{'=' * 60}")
    records = extract_invoice_fields(
        groups, manifest, client,
        include_ocr_text=include_ocr_text,
        log_callback=log_callback,
        progress_callback=progress_callback,
        cancel_event=cancel_event,
    )
    if cancel_event and cancel_event.is_set():
        return None

    # Write outputs
    log(f"\n{'=' * 60}")
    log("OUTPUT")
    log(f"{'=' * 60}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    csv_path = os.path.join(output_dir, f"invoices_{timestamp}.csv")
    write_csv(records, csv_path)
    log(f"CSV  → {csv_path}")

    xlsx_path = os.path.join(output_dir, f"invoices_{timestamp}.xlsx")
    write_xlsx(records, xlsx_path)
    log(f"XLSX → {xlsx_path}")

    json_path = os.path.join(output_dir, f"invoices_{timestamp}.json")
    with open(json_path, "w") as f:
        json.dump([asdict(r) for r in records], f, indent=2, default=str)
    log(f"JSON → {json_path}")

    invoice_count = sum(1 for r in records if r.group_type == "invoice")
    support_count = sum(1 for r in records if r.group_type != "invoice")
    high_conf = sum(1 for r in records if r.group_type == "invoice" and r.extraction_confidence >= 0.7)
    low_conf = sum(1 for r in records if r.group_type == "invoice" and r.extraction_confidence < 0.5)

    log(f"\n{'=' * 60}")
    log(f"SUMMARY")
    log(f"  Invoices extracted: {invoice_count}")
    log(f"  Supporting docs:    {support_count}")
    log(f"  High confidence:    {high_conf}")
    log(f"  Low confidence:     {low_conf} (review recommended)")
    log(f"{'=' * 60}")

    return xlsx_path


# ===========================================================================
# GUI — Base tab with shared widgets
# ===========================================================================
class BaseTab(ttk.Frame):
    """Shared GUI skeleton: progress bar, log text, start/cancel/open buttons."""

    def __init__(self, parent, start_label="Start"):
        super().__init__(parent, padding=10)
        self.processing = False
        self.cancel_event = threading.Event()
        self._start_label = start_label
        self._build_base_widgets()

    def _build_base_widgets(self):
        """Create the common bottom section: progress, buttons, log."""
        # Subclasses add their own widgets first, then call _build_base_widgets
        # via _finish_layout() after adding custom controls.
        pass

    def _finish_layout(self):
        """Call this at the end of subclass _build_ui to add progress/buttons/log."""
        # --- Progress ---
        prog_frame = ttk.LabelFrame(self, text="Progress", padding=5)
        prog_frame.pack(fill=tk.X, pady=(0, 5))
        self.progress_label = ttk.Label(prog_frame, text="Ready")
        self.progress_label.pack(anchor=tk.W)
        self.progress_bar = ttk.Progressbar(prog_frame, mode="determinate")
        self.progress_bar.pack(fill=tk.X, pady=(3, 0))

        # --- Buttons ---
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=(5, 5))
        self.start_btn = ttk.Button(btn_frame, text=f"▶  {self._start_label}",
                                    command=self._start_processing)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.cancel_btn = ttk.Button(btn_frame, text="■  Cancel",
                                     command=self._cancel, state=tk.DISABLED)
        self.cancel_btn.pack(side=tk.LEFT, padx=(0, 5))
        self.open_btn = ttk.Button(btn_frame, text="Open Output",
                                   command=self._open_output, state=tk.DISABLED)
        self.open_btn.pack(side=tk.RIGHT)

        # --- Log ---
        log_frame = ttk.LabelFrame(self, text="Processing Log", padding=5)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, wrap=tk.WORD,
                                                   font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _start_processing(self):
        """Override in subclass."""
        pass

    def _cancel(self):
        self.cancel_event.set()
        self._append_log("Cancelling...")

    def _open_output(self):
        """Override in subclass to provide the folder path."""
        pass

    def _set_running(self, running: bool):
        self.processing = running
        self.start_btn.config(state=tk.DISABLED if running else tk.NORMAL)
        self.cancel_btn.config(state=tk.NORMAL if running else tk.DISABLED)
        if not running:
            self.open_btn.config(state=tk.NORMAL)

    def _update_progress(self, current, total, msg):
        def _u():
            pct = (current / total * 100) if total > 0 else 0
            self.progress_bar["value"] = pct
            self.progress_label.config(text=f"[{current}/{total}] {msg}")
        self.winfo_toplevel().after(0, _u)

    def _append_log(self, message):
        def _a():
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
        self.winfo_toplevel().after(0, _a)


# ===========================================================================
# GUI — Phase 1 Tab
# ===========================================================================
class Phase1Tab(BaseTab):
    def __init__(self, parent, app):
        self.app = app
        super().__init__(parent, start_label="Start Processing")
        self._build_ui()

    def _build_ui(self):
        self.input_folder = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.recursive = tk.BooleanVar(value=False)

        # --- Dependency status ---
        dep_frame = ttk.LabelFrame(self, text="System Dependencies", padding=5)
        dep_frame.pack(fill=tk.X, pady=(0, 10))
        self.dep_label = ttk.Label(dep_frame, text="Checking...")
        self.dep_label.pack(anchor=tk.W)
        self._check_dependencies()

        # --- Input folder ---
        input_frame = ttk.LabelFrame(self, text="Input Folder (containing PDFs)", padding=5)
        input_frame.pack(fill=tk.X, pady=(0, 5))
        input_row = ttk.Frame(input_frame)
        input_row.pack(fill=tk.X)
        ttk.Entry(input_row, textvariable=self.input_folder).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(input_row, text="Browse…", command=self._browse_input).pack(side=tk.RIGHT)
        ttk.Checkbutton(input_frame, text="Include subfolders (recursive)",
                        variable=self.recursive).pack(anchor=tk.W, pady=(3, 0))

        # --- Output folder ---
        output_frame = ttk.LabelFrame(self, text="Output Folder", padding=5)
        output_frame.pack(fill=tk.X, pady=(0, 5))
        output_row = ttk.Frame(output_frame)
        output_row.pack(fill=tk.X)
        ttk.Entry(output_row, textvariable=self.output_folder).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_row, text="Browse…", command=self._browse_output).pack(side=tk.RIGHT)

        self._finish_layout()

    def _check_dependencies(self):
        missing = check_system_dependencies()
        if missing:
            tool_names = {"tesseract": "tesseract-ocr", "pdftoppm": "poppler-utils", "gs": "ghostscript"}
            packages = [tool_names.get(t, t) for t in missing]
            if sys.platform == "win32":
                install_hint = "Download from: https://github.com/UB-Mannheim/tesseract/wiki (tesseract), https://github.com/oschwartz10612/poppler-windows (poppler), https://www.ghostscript.com/download.html (ghostscript)"
            else:
                install_hint = f"Install with: sudo apt install {' '.join(packages)}"
            self.dep_label.config(
                text=f"⚠ Missing: {', '.join(packages)}. {install_hint}",
                foreground="red",
            )
        else:
            self.dep_label.config(text="✓ All dependencies found (tesseract, poppler, ghostscript)",
                                  foreground="green")

    def _browse_input(self):
        folder = filedialog.askdirectory(title="Select folder containing PDFs")
        if folder:
            self.input_folder.set(folder)
            if not self.output_folder.get():
                self.output_folder.set(os.path.join(folder, "processed_output"))

    def _browse_output(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self.output_folder.set(folder)

    def _start_processing(self):
        input_dir = self.input_folder.get().strip()
        output_dir = self.output_folder.get().strip()

        if not input_dir or not os.path.isdir(input_dir):
            messagebox.showerror("Error", "Please select a valid input folder.")
            return
        if not output_dir:
            messagebox.showerror("Error", "Please select an output folder.")
            return

        os.makedirs(output_dir, exist_ok=True)
        self.log_text.delete("1.0", tk.END)
        self.progress_bar["value"] = 0
        self.cancel_event.clear()
        self._set_running(True)

        thread = threading.Thread(target=self._run_thread,
                                  args=(input_dir, output_dir), daemon=True)
        thread.start()

    def _run_thread(self, input_dir, output_dir):
        try:
            records = run_phase1(
                input_folder=input_dir,
                output_folder=output_dir,
                recursive=self.recursive.get(),
                progress_callback=self._update_progress,
                log_callback=self._append_log,
                cancel_event=self.cancel_event,
            )
            self.winfo_toplevel().after(0, self._done, len(records), output_dir)
        except Exception as e:
            self.winfo_toplevel().after(0, self._error, str(e))

    def _done(self, count, output_dir):
        self._set_running(False)
        self.progress_label.config(text=f"Complete — {count} pages processed")
        self.progress_bar["value"] = 100

        # Auto-populate Phase 2 manifest path
        manifest_path = os.path.join(output_dir, "manifest.json")
        if os.path.isfile(manifest_path):
            self.app.phase2_tab.manifest_path.set(manifest_path)
            if not self.app.phase2_tab.output_folder.get():
                self.app.phase2_tab.output_folder.set(
                    os.path.join(output_dir, "extraction_output"))
            self._append_log(f"\nPhase 2 manifest auto-set → switch to Extraction tab to continue")

    def _error(self, msg):
        self._set_running(False)
        messagebox.showerror("Processing Error", f"An error occurred:\n{msg}")

    def _open_output(self):
        open_folder(self.output_folder.get())


# ===========================================================================
# GUI — Phase 2 Tab
# ===========================================================================
class Phase2Tab(BaseTab):
    def __init__(self, parent, app):
        self.app = app
        super().__init__(parent, start_label="Start Extraction")
        self._build_ui()

    def _build_ui(self):
        self.manifest_path = tk.StringVar()
        self.output_folder = tk.StringVar()

        # --- Input ---
        input_frame = ttk.LabelFrame(self, text="Phase 1 Output (manifest.json)", padding=5)
        input_frame.pack(fill=tk.X, pady=(0, 5))
        input_row = ttk.Frame(input_frame)
        input_row.pack(fill=tk.X)
        ttk.Entry(input_row, textvariable=self.manifest_path).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(input_row, text="Browse…", command=self._browse_manifest).pack(side=tk.RIGHT)

        # --- Output ---
        output_frame = ttk.LabelFrame(self, text="Output Folder", padding=5)
        output_frame.pack(fill=tk.X, pady=(0, 5))
        output_row = ttk.Frame(output_frame)
        output_row.pack(fill=tk.X)
        ttk.Entry(output_row, textvariable=self.output_folder).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_row, text="Browse…", command=self._browse_output).pack(side=tk.RIGHT)

        self._finish_layout()

    def _browse_manifest(self):
        path = filedialog.askopenfilename(
            title="Select manifest.json from Phase 1",
            filetypes=[("JSON", "*.json"), ("All", "*.*")],
        )
        if path:
            self.manifest_path.set(path)
            if not self.output_folder.get():
                self.output_folder.set(
                    os.path.join(os.path.dirname(path), "extraction_output"))

    def _browse_output(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self.output_folder.set(folder)

    def _start_processing(self):
        manifest = self.manifest_path.get().strip()
        output = self.output_folder.get().strip()

        if not manifest or not os.path.isfile(manifest):
            messagebox.showerror("Error", "Select a valid manifest.json.")
            return
        if not output:
            messagebox.showerror("Error", "Select an output folder.")
            return

        self.log_text.delete("1.0", tk.END)
        self.progress_bar["value"] = 0
        self.cancel_event.clear()
        self._set_running(True)

        thread = threading.Thread(target=self._run_thread,
                                  args=(manifest, output), daemon=True)
        thread.start()

    def _run_thread(self, manifest_path, output_dir):
        try:
            result = run_phase2(
                manifest_path=manifest_path,
                output_dir=output_dir,
                ollama_url=self.app.settings_tab.ollama_url.get().strip(),
                model=self.app.settings_tab.model_name.get().strip(),
                include_ocr_text=self.app.settings_tab.include_ocr.get(),
                log_callback=self._append_log,
                progress_callback=self._update_progress,
                cancel_event=self.cancel_event,
            )
            self.winfo_toplevel().after(0, self._done, result)
        except Exception as e:
            self.winfo_toplevel().after(0, self._error, str(e))

    def _done(self, result_path):
        self._set_running(False)
        self.progress_bar["value"] = 100
        if result_path:
            self.progress_label.config(text=f"Complete → {os.path.basename(result_path)}")
        else:
            self.progress_label.config(text="Completed with errors — check log")

    def _error(self, msg):
        self._set_running(False)
        messagebox.showerror("Error", msg)

    def _open_output(self):
        open_folder(self.output_folder.get())


# ===========================================================================
# GUI — Settings Tab
# ===========================================================================
class SettingsTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=10)
        self.ollama_url = tk.StringVar(value="http://192.168.1.101:11434")
        self.model_name = tk.StringVar(value="qwen3-vl:8b")
        self.include_ocr = tk.BooleanVar(value=True)
        self._build_ui()

    def _build_ui(self):
        # --- Ollama settings ---
        ollama_frame = ttk.LabelFrame(self, text="Ollama Settings", padding=10)
        ollama_frame.pack(fill=tk.X, pady=(0, 10))

        row1 = ttk.Frame(ollama_frame)
        row1.pack(fill=tk.X, pady=4)
        ttk.Label(row1, text="Server URL:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(row1, textvariable=self.ollama_url, width=35).pack(side=tk.LEFT, padx=(0, 10))

        row2 = ttk.Frame(ollama_frame)
        row2.pack(fill=tk.X, pady=4)
        ttk.Label(row2, text="Vision Model:").pack(side=tk.LEFT, padx=(0, 5))
        ttk.Entry(row2, textvariable=self.model_name, width=25).pack(side=tk.LEFT, padx=(0, 10))

        row3 = ttk.Frame(ollama_frame)
        row3.pack(fill=tk.X, pady=4)
        self.ollama_status = ttk.Label(row3, text="")
        self.ollama_status.pack(side=tk.LEFT)
        ttk.Button(row3, text="Test Connection", command=self._test_ollama).pack(side=tk.RIGHT)

        # --- OCR settings ---
        ocr_frame = ttk.LabelFrame(self, text="Extraction Settings", padding=10)
        ocr_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Checkbutton(ocr_frame,
                        text="Include OCR text as supplementary context (recommended)",
                        variable=self.include_ocr).pack(anchor=tk.W)

    def _test_ollama(self):
        url = self.ollama_url.get().strip()
        model = self.model_name.get().strip()
        client = OllamaClient(base_url=url, model=model)

        if client.is_available():
            self.ollama_status.config(text="✓ Connected — model ready", foreground="green")
        else:
            models = client.list_models()
            if models:
                self.ollama_status.config(
                    text=f"⚠ Model not found. Available: {', '.join(models[:5])}",
                    foreground="orange")
            else:
                self.ollama_status.config(text="✗ Cannot reach Ollama", foreground="red")


# ===========================================================================
# GUI — Main Application
# ===========================================================================
class InvoiceProcessorApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("PDF Invoice Processor")
        self.root.geometry("860x750")
        self.root.minsize(720, 600)

        # Apply theme
        try:
            style = ttk.Style()
            for t in ("clam", "alt", "vista", "xpnative"):
                if t in style.theme_names():
                    style.theme_use(t)
                    break
        except Exception:
            pass

        # Header
        header = ttk.Label(root, text="PDF Invoice Processor",
                           font=("Segoe UI", 16, "bold"))
        header.pack(pady=(10, 2))
        subtitle = ttk.Label(root,
                             text="Split · Orient · OCR · Classify · Extract — 100% Local",
                             font=("Segoe UI", 10))
        subtitle.pack(pady=(0, 8))

        # Notebook (tabs)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # Settings tab first (so other tabs can reference it)
        self.settings_tab = SettingsTab(self.notebook)

        # Phase tabs
        self.phase1_tab = Phase1Tab(self.notebook, self)
        self.phase2_tab = Phase2Tab(self.notebook, self)

        self.notebook.add(self.phase1_tab, text="  1. Document Preparation  ")
        self.notebook.add(self.phase2_tab, text="  2. Vision Extraction  ")
        self.notebook.add(self.settings_tab, text="  Settings  ")


# ===========================================================================
# Entry point
# ===========================================================================
def main():
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
    root = tk.Tk()
    InvoiceProcessorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
